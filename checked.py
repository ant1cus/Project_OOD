import os
import re
import pathlib

import psutil
from openpyxl import load_workbook


def check(n, e):
    for el in e:
        if n == el:
            return False
    return True


def checked_zone_checked(line_edit_path_check, line_edit_table_number, zone):

    for proc in psutil.process_iter():
        if proc.name() == 'WINWORD.EXE':
            return ['УПС!', 'Закройте все файлы Word!']
    path = line_edit_path_check.text().strip()
    if not path:
        return ['УПС!', 'Путь к проверяемым документам пуст']
    if not os.path.isdir(path):
        return ['УПС!', 'Указанный путь к проверяемым документам не является директорией']
    if not os.listdir(path):
        return ['УПС!', 'В указанной директории отсутствуют файлы для проверки']
    table = line_edit_table_number.text().strip()
    if not table:
        return ['УПС!', 'Не указан номер таблицы']
    for i in table:
        if check(i, ('1', '2', '3', '4', '5', '6', '7', '8', '9', '0')):
            return ['УПС!', 'Есть лишние символы в номере таблицы']
    zone_out = {i: el.text().replace(',', '.') if el.text() else '10000000000' for i, el in enumerate(zone)}
    err_msg = ('Есть лишние символы в ограничении по стационарным антеннам',
               'Есть лишние символы в ограничении по возимым антеннам',
               'Есть лишние символы в ограничении по носимым антеннам',
               'Есть лишние символы в ограничении r1',
               'Есть лишние символы в ограничении r1`')
    if zone_out:
        for i in zone_out:
            for j in zone_out[i]:
                if check(j, ('1', '2', '3', '4', '5', '6', '7', '8', '9', '0', '.')):
                    return ['УПС!', err_msg[i]]
    else:
        return ['УПС!', 'Не указано ни одно ограничение для проверки']
    return zone_out


def checked_file_parcing(dir_path, group_file):
    def folder_checked(p):
        errors = []
        txt_files = list(filter(lambda x: x.endswith('.txt'), os.listdir(p)))
        excel_files = [x for x in os.listdir(p) if x.endswith('.xlsx')]
        if 'Описание.txt' not in txt_files and 'описание.txt' not in txt_files:
            errors.append('Нет файла с описанием режимов (' + p + ')')
        else:
            try:
                # with open(path + '\\' + file, mode='r', encoding="utf-8-sig") as f:
                with open(pathlib.Path(p, 'Описание.txt'), mode='r', encoding='utf-8') as f:
                    lines = f.readlines()
            except UnicodeDecodeError:
                # with open(path + '\\' + file, mode='r') as f:
                with open(pathlib.Path(p, 'Описание.txt'), mode='r', encoding='ANSI') as f:
                    lines = f.readlines()
            for line in lines:
                if re.findall(r'\s', line.rstrip('\n')):
                    errors.append('Пробелы в названии режимов (' + p + ', ' + line.rstrip('\n') + ')')
            # with open(pathlib.Path(p, 'Описание.txt'), mode='r', encoding='utf-8') as f:
            #     lines = f.readlines()
            #     for line in lines:
            #         if re.findall(r'\s', line.rstrip('\n')):
            #             errors.append('Пробелы в названии режимов (' + p + ', ' + line.rstrip('\n') + ')')
        return {'errors': errors, 'len': len(excel_files)}
    # Выбираем путь для исходников.
    path = dir_path.text().strip()
    if not path:
        return ['УПС!', 'Не указан путь к исходной папке']
    elif not os.path.isdir(path):
        return ['УПС!', 'Путь к исходной папке удалён или переименован']
    else:
        folders = [i for i in os.listdir(path) if os.path.isdir(path + '\\' + i) and i != 'txt']
        if group_file is False and folders:
            return ['УПС!', 'В директории для парсинга присутствуют папки']
        elif group_file and folders is False:
            return ['УПС!', 'В директории для парсинга нет папок для преобразования']
    error = []
    progress = 0
    if group_file:
        for folder in os.listdir(path):
            if os.path.isdir(path + '\\' + folder):
                err = folder_checked(path + '\\' + folder)
                progress += err['len']
                if err['errors']:
                    error.append(err)
    else:
        err = folder_checked(path)
        error, progress = err['errors'], err['len']
    return ['УПС!', '\n'.join(error)] if error else {'path': path, 'progress': progress}


def checked_generation_pemi(source_file, output_file, set_number, set_quantity, freq_restrict,
                            freq_restrict_path):

    source = source_file.text().strip()
    if not source:
        return ['УПС!', 'Путь к исходным файлам пуст']
    if not os.path.isdir(source):
        return ['УПС!', 'Путь к исходным файлам удалён или переименован']
    output = output_file.text().strip()
    if not output:
        return ['УПС!', 'Путь к создаваемым файлам пуст']
    if not os.path.isdir(output):
        return ['УПС!', 'Путь к создаваемым файлам удалён или переименован']
    if len([True for el in pathlib.Path(source).iterdir() if el.is_dir()]) > 1:
        return ['УПС!', 'В указанной директории слишком много папок']
    set_num = set_number.text().strip()
    if not set_num:
        return ['УПС!', 'Не указаны номера комплектов']
    for i in set_num:
        if check(i, ('1', '2', '3', '4', '5', '6', '7', '8', '9', '0', ' ', '-', ',', '.')):
            return ['УПС!', 'Есть лишние символы в номерах комплектов']
    set_num = set_num.replace(' ', '').replace(',', '.')
    if set_num[0] == '.' or set_num[0] == '-':
        return ['УПС!', 'Первый символ введён не верно']
    if set_num[-1] == '.' or set_num[-1] == '-':
        return ['УПС!', 'Последний символ введён не верно']
    for i in range(len(set_num)):
        if set_num[i] == '.' or set_num[i] == '-':
            if set_num[i + 1] == '.' or set_num[i + 1] == '-':
                return ['УПС!', 'Два разделителя номеров подряд']
    set_list = []
    for element in set_num.split('.'):
        if '-' in element:
            num1, num2 = int(element.partition('-')[0]), int(element.partition('-')[2])
            if num1 >= num2:
                return ['УПС!', 'Диапазон номеров комплектов указан не верно']
            else:
                for el in range(num1, num2 + 1):
                    set_list.append(el)
        else:
            set_list.append(element)
    set_list.sort()
    if len(set_list) != len(set(set_list)):
        return ['УПС!', 'Есть повторения в номерах комплектов']
    set_quant = set_quantity.text()
    if not set_quant:
        return ['УПС!', 'Не указано количество комплектов']
    for i in set_quant:
        if check(i, ('1', '2', '3', '4', '5', '6', '7', '8', '9', '0')):
            return ['УПС!', 'Не правильно указано количество комплектов']
    if len(set_list) != int(set_quant):
        return ['УПС!', 'Указанные номера не совпадают с количеством генерируемых комплектов']
    txt_files = list(filter(lambda x: x.endswith('.txt'), os.listdir(source)))
    if len(txt_files) != 1:
        return ['УПС!', 'В папке больше одного или нет txt файла']
    name_mode = ''
    for file in sorted(txt_files):
        if os.stat(source + '\\' + file).st_size == 0:
            return ['УПС!', 'Файл с описанием режимов пуст']
        else:
            try:
                with open(str(pathlib.Path(source, file)), mode='r', encoding='utf-8') as f:
                    name_mode = f.readlines()
                    if ' ' in name_mode:
                        return ['УПС!', 'Пробелы в названии режимов в файле «Описание»']
                    else:
                        name_mode = [line.rstrip().lower() for line in name_mode if line]
            except UnicodeDecodeError:
                with open(str(pathlib.Path(source, file)), mode='r') as f:
                    name_mode = f.readlines()
                    if ' ' in name_mode:
                        return ['УПС!', 'Пробелы в названии режимов в файле «Описание»']
                    else:
                        name_mode = [line.rstrip().lower() for line in name_mode if line]
    error = []
    for file_exel in sorted(list(filter(lambda x: x.endswith('.xlsx'), os.listdir(output)))):
        if int(file_exel[:-5]) in set_list:
            error.append(file_exel)
    if error:
        return ['УПС!', 'В указанной папке уже есть такие исходники:\n' + '\n'.join(error)]
    error = []
    for file_exel in sorted(list(filter(lambda x: x.endswith('.xlsx'), os.listdir(source)))):
        wb = load_workbook(str(pathlib.Path(source, file_exel)))  # Откроем книгу.
        name = wb.sheetnames  # Список листов.
        for name_exel in name:
            if name_exel.lower() not in name_mode:
                error.append('Названия режимов в файле ' + file_exel + ' не совпадают с описанием.')
                break
        wb.close()
    if error:
        return ['УПС!', '\n'.join(error)]
    # Добавление для разницы!
    restrict_file = False
    if freq_restrict:
        restrict_file = freq_restrict_path.text().strip()
        if not restrict_file:
            return ['УПС!', 'Путь к файлу с ограничениями пуст']
        if os.path.isfile(restrict_file):
            if restrict_file.endswith('.txt'):
                if os.stat(restrict_file).st_size == 0:
                    return ['УПС!', 'Файл для комплектов пуст']
            else:
                return ['УПС!', 'Файл для комплектов не txt']
        else:
            return ['УПС!', 'Указана директория в файле для ограничений']
    return {'source': source, 'output': output, 'set': set_list, 'set_quant': set_quant,
            'name_mode': name_mode, 'restrict_file': restrict_file}


def checked_delete_header_footer(path, check_box_director, line_edit_old_director, line_edit_new_director):
    source = path.text().strip()
    if not source:
        return ['УПС!', 'Путь к исходным файлам пуст']
    if not os.path.isdir(source):
        return ['УПС!', 'Путь к исходным файлам удалён или переименован']
    if not os.listdir(source):
        return ['УПС!', 'В указанной директории отсутствуют файлы для обезличивания']
    name_file = [False, False, False]
    error = []
    files = [file for file in os.listdir(source) if '~' not in file and file.endswith('.docx')]
    if not files:
        return ['УПС!', 'В указанной директории отсутствуют файлы, пригодные для обезличивания']
    for file in files:
        if file.endswith('.doc'):
            error.append(file)
        if 'заключение' in file.lower():
            name_file[0] = True
        elif 'протокол' in file.lower():
            name_file[1] = True
        elif 'предписание' in file.lower():
            name_file[2] = True
    if error:
        return ['УПС!', 'Файлы старого формата:\n' + '\n'.join(error)]
    old_director = None
    new_director = None
    if check_box_director.isChecked():
        old_director = line_edit_old_director.text()
        new_director = line_edit_new_director.text()
    return {'path': source, 'old_director': old_director, 'new_director': new_director}


def checked_hfe_generation(path, set_value, req_val, frequency, value):
    output = path.text().strip()
    if not output:
        return ['УПС!', 'Не указан путь к конечной папке']
    if not os.path.isdir(output):
        return ['УПС!', 'Путь к конечной папке удалён или переименован']
    if req_val.isChecked():
        freq = frequency.text().strip()
        val = value.text().strip()
    else:
        freq = '100'
        val = '100'
    quantity = set_value.text().strip()
    variables = [freq, val, quantity]
    errors1 = ['Не указана частота', 'Не указан уровень', 'Не указано количество комплектов']
    errors2 = ['Частота указана с ошибкой', 'Уровень указан с ошибкой', 'Количество комплектов указано с ошибкой']
    for i in range(0, 3, 1):
        if not variables[i]:
            return ['УПС!', errors1[i]]
        for j in variables[i]:
            if check(j, ('1', '2', '3', '4', '5', '6', '7', '8', '9', '0')):
                return ['УПС!', errors2[i]]
    return {'path': output, 'quantity': int(quantity), 'freq': freq, 'val': val}


def checked_hfi_generation(path, frequency, quantity, mode):
    output = path.text().strip()
    if not output:
        return ['УПС!', 'Не указан путь к конечной папке']
    if not os.path.isdir(output):
        return ['УПС!', 'Путь к конечной папке удалён или переименован']
    variables = [frequency.text().strip(), quantity.text().strip()]
    errors1 = ['Не указана частота навязывания', 'Не указано количество комплектов']
    errors2 = ['Частота навязывания указана с ошибкой', 'Количество комплектов указано с ошибкой']
    for i in range(0, 2, 1):
        if not variables[i]:
            return ['УПС!', errors1[i]]
        for j in variables[i]:
            if check(j, ('1', '2', '3', '4', '5', '6', '7', '8', '9', '0')):
                return ['УПС!', errors2[i]]
    if any(mode):
        return {'path': output, 'quantity': int(quantity.text().strip()), 'freq': variables[0], 'val': variables[1],
                'mode': mode}
    else:
        return ['УПС!', 'Не выбран не один режим']


def checked_application_data(file_path, folder_path, number, quant):
    file = file_path.text().strip()
    if not file:
        return ['УПС!', 'Не указан путь к файлу для генерации']
    if os.path.isfile(file):
        if not file.endswith('.docx'):
            return ['УПС!', 'Указанный файл неверного формата (необходим docx)']
    else:
        return ['УПС!', 'Указанный файл удалён или переименован']
    folder = folder_path.text().strip()
    if not folder:
        return ['УПС!', 'Не указан путь к конечной папке для генерации']
    if not os.path.isdir(folder):
        return ['УПС!', 'Путь к конечной папке удалён или переименован']
    position_num = number.text().strip()
    if not position_num:
        return ['УПС!', 'Не указан номер позиции']
    for i in position_num:
        if check(i, ('1', '2', '3', '4', '5', '6', '7', '8', '9', '0')):
            return ['УПС!', 'Не правильно указан номер позиции']
    quantity = quant.text().strip()
    if not quantity:
        return ['УПС!', 'Не указано количество комплектов']
    for i in quantity:
        if check(i, ('1', '2', '3', '4', '5', '6', '7', '8', '9', '0')):
            return ['УПС!', 'Не правильно указано количество комплектов']
    return {'file': file, 'path': folder, 'quantity': int(quantity), 'position_num': int(position_num)}


def checked_lf_data(source_folder, output_folder, excel_file):
    source = source_folder.text().strip()
    if not source:
        return ['УПС!', 'Путь к исходным файлам пуст']
    if not os.path.isdir(source):
        return ['УПС!', 'Указанный путь к исходным файлам удалён или переименован']
    output = output_folder.text().strip()
    if not output:
        return ['УПС!', 'Путь к создаваемым файлам пуст']
    if not os.path.isdir(output):
        return ['УПС!', 'Путь к создаваемым файлам удалён или переименован']
    file = excel_file.text().strip()
    if not file:
        return ['УПС!', 'Путь к файлу генератору пуст']
    else:
        if file.endswith('.xlsx'):
            return {'source': source, 'output': output, 'excel': file}
        else:
            return ['УПС!', 'Указанный файл недопустимого формата (необходимо .xlsx)']


def checked_generation_cc(start_folder, finish_folder, set_number, checkbox_frequency, frequency, checkbox_txt,
                          checkbox_dispersion, lineedit_dispersion):

    source = start_folder.text().strip()
    if not source:
        return ['УПС!', 'Путь к исходным файлам пуст']
    if not os.path.isdir(source):
        return ['УПС!', 'Указанный путь к исходным файлам переименован или удалён']
    output = finish_folder.text().strip()
    if not output:
        return ['УПС!', 'Путь к создаваемым файлам пуст']
    if not os.path.isdir(output):
        return ['УПС!', 'Указанный путь к создаваемым файлам переименован или удалён']
    if len([True for el in pathlib.Path(source).iterdir() if el.is_dir()]) > 1:
            return ['УПС!', 'В указанной директории слишком много папок']
    set_num = set_number.text().strip()
    if set_num:
        for i in set_num:
            if check(i, ('1', '2', '3', '4', '5', '6', '7', '8', '9', '0', ' ', '-', ',', '.')):
                return ['УПС!', 'Есть лишние символы в номерах комплектов']
        set_num = set_num.replace(' ', '').replace(',', '.')
        if set_num[0] == '.' or set_num[0] == '-':
            return ['УПС!', 'Первый символ введён не верно']
        if set_num[-1] == '.' or set_num[-1] == '-':
            return ['УПС!', 'Последний символ введён не верно']
        for i in range(len(set_num)):
            if set_num[i] == '.' or set_num[i] == '-':
                if set_num[i + 1] == '.' or set_num[i + 1] == '-':
                    return ['УПС!', 'Два разделителя номеров подряд']
        set_list = []
        for element in set_num.split('.'):
            if '-' in element:
                num1, num2 = int(element.partition('-')[0]), int(element.partition('-')[2])
                if num1 >= num2:
                    return ['УПС!', 'Диапазон номеров комплектов указан не верно']
                else:
                    for el in range(num1, num2 + 1):
                        set_list.append(el)
            else:
                set_list.append(element)
        set_list.sort()
        if len(set_list) != len(set(set_list)):
            return ['УПС!', 'Есть повторения в номерах комплектов']
        if len(set_list) == 1 and set_list[0] == '0':
            set_num = False
        else:
            set_num = set_list
    freq = frequency.text() if checkbox_frequency.isChecked() else ''
    if freq:
        for i in freq:
            if check(i, ('1', '2', '3', '4', '5', '6', '7', '8', '9', '0', '.')):
                return ['УПС!', 'Не правильно указана частота для ограничения (дробный разделитель - ".")']
    txt = checkbox_txt.isChecked()
    dispersion = lineedit_dispersion.text() if checkbox_dispersion.isChecked() else ''
    if dispersion:
        for i in dispersion:
            if check(i, ('1', '2', '3', '4', '5', '6', '7', '8', '9', '0', '.')):
                return ['УПС!', 'Не правильно указан разброс (только цифры, дробный разделитель - ".")']
    return {'source': source, 'output': output, 'set': set_num, 'frequency': freq, 'only_txt': txt,
            "dispersion": float(dispersion)}


def checked_number_instance(start_folder, finish_folder, incoming_set_number):
    path_old = start_folder.text()
    if not path_old:
        return ['УПС!', 'Путь к исходным файлам пуст']
    if os.path.isdir(path_old):
        pass
    else:
        return ['УПС!', 'Указанный путь к исходным файлам не является директорией']
    path_new = finish_folder.text()
    if not path_new:
        return ['УПС!', 'УПС!', 'Путь к конечным файлам пуст']
    if os.path.isdir(path_new):
        pass
    else:
        return ['УПС!', 'Указанный путь к конечным файлам не является директорией']
    number_instance_ = incoming_set_number.text().strip()
    for i in number_instance_:
        if check(i, ('1', '2', '3', '4', '5', '6', '7', '8', '9', '0', ' ', '-', ',', '.')):
            return ['УПС!', 'Есть лишние символы в номерах экземпляров']
    set_num = number_instance_.replace(' ', '').replace(',', '.')
    if set_num[0] == '.' or set_num[0] == '-':
        return ['УПС!', 'Первый символ введён не верно']
    if set_num[-1] == '.' or set_num[-1] == '-':
        return ['УПС!', 'Последний символ введён не верно']
    for i in range(len(set_num)):
        if set_num[i] == '.' or set_num[i] == '-':
            if set_num[i + 1] == '.' or set_num[i + 1] == '-':
                return ['УПС!', 'Два разделителя номеров подряд']
    set_number = []
    for element in set_num.split('.'):
        if '-' in element:
            num1, num2 = int(element.partition('-')[0]), int(element.partition('-')[2])
            if num1 >= num2:
                return ['УПС!', 'Диапазон номеров экземпляров указан не верно']
            else:
                for el in range(num1, num2 + 1):
                    set_number.append(el)
        else:
            set_number.append(element)
    set_number.sort()
    return {'path_old': path_old, 'path_new': path_new, 'set_number': set_number}


def checked_find_files(unloading_file, start_folder, finish_folder):
    file = unloading_file.text()
    if not file:
        return ['УПС!', 'Поле «файл с выгрузкой» пусто']
    if os.path.isdir(file):
        return ['УПС!', 'В поле «файл с выгрузкой» указана директория, а не файл']
    if file.endswith('.xlsx') is False and file.endswith('.txt') is False:
        return ['УПС!', 'Указанный файл с выгрузкой не требуемого формата (необходим ".xlsx" или ".txt")']
    start_path = start_folder.text()
    if not start_path:
        return ['УПС!', 'Путь к исходным файлам пуст']
    if os.path.isdir(start_path) is False:
        return ['УПС!', 'Указанный путь к исходным файлам не является директорией']
    finish_path = finish_folder.text()
    if not finish_path:
        return ['УПС!', 'УПС!', 'Путь к конечной папке пуст']
    if os.path.isdir(finish_path) is False:
        return ['УПС!', 'Указанный путь к конечной папке не является директорией']
    if os.listdir(finish_path):
        return ['УПС!', 'Конечная папка не пуста, очистите директорию или выберите новую']

    return {'unloading_file': file, 'start_path': start_path, 'finish_path': finish_path}

import datetime
import os
import re
import traceback
import pathlib

import pandas as pd
from openpyxl import load_workbook


def file_parcing(path, logging, line_doing, now_doc, all_doc, line_progress, progress, per, cp, no_freq_lim,
                 default_path, event, window_check, twelve_sectors=False):
    try:
        list_file = os.listdir(path)
        # Сохраним нужное нам описание режимов.
        logging.info("Читаем txt и сохраняем режимы для " + path)
        txt_files = filter(lambda x: x.endswith('.txt'), list_file)
        for file in sorted(txt_files):
            try:
                # with open(path + '\\' + file, mode='r', encoding="utf-8-sig") as f:
                with open(pathlib.Path(path, file), mode='r', encoding="utf-8-sig") as f:
                    logging.info("Кодировка utf-8-sig")
                    mode_1 = f.readlines()
                    mode_1 = [line.rstrip() for line in mode_1]
            except UnicodeDecodeError:
                # with open(path + '\\' + file, mode='r') as f:
                with open(pathlib.Path(path, file), mode='r') as f:
                    logging.info("Другая кодировка")
                    mode_1 = f.readlines()
                    mode_1 = [line.rstrip() for line in mode_1]
        mode = [x for x in mode_1 if x]
        parcing_file = []
        if os.path.exists(pathlib.Path(path, 'txt')):
            logging.info("Запоминаем какие папки уже есть внутри папки txt")
            parcing_file = os.listdir(pathlib.Path(path, 'txt'))
        # Работа с исходниками.
        # Отсортируем нужные нам файлы xlsx.
        exel_files = filter(lambda x: x.endswith('.xlsx') and ('~' not in x) and (x[:-4] not in parcing_file),
                            list_file)
        logging.info("Начинаем прохождение по файлам excel")
        output_error = []
        for file in sorted(exel_files):
            event.wait()
            if window_check.stop_threading:
                return {'cancel': True}
            now_doc += 1
            line_doing.emit(f'Проверяем названия рабочих листов в документе {file} ({now_doc} из {all_doc})')
            # status.emit('Проверяем названия рабочих листов в документе ' + file)
            error = []
            pat = ['_ЦП', '.m', '.v']  # список ключевых слов для поиска в ЦП
            pat_rez = ['_ЦП', '.m', '.v']
            book_name = file.rpartition('.xlsx')[0]
            logging.info("Открываем книгу")
            # Новое для ускорения
            new_book = {}
            book = pd.read_excel(pathlib.Path(path, file), sheet_name=None, header=None)
            for enum, name_list in enumerate(book.keys()):
                x = name_list
                if re.search(r'_ЦП', name_list) or re.search(r'\.m', name_list) or re.search(r'\.v', name_list):
                    logging.info(f"Нашли название {name_list}")
                    rez = []
                    for y in pat:  # прогоняем список
                        logging.info("Ищем совпадение в нашем списке")
                        if y == '.v':
                            replace = re.findall(r'.v\d', x)
                            if replace:
                                pat_rez[2] = replace[0]
                        rez.append(1) if x.find(y) != -1 else rez.append(-1)  # добавляем заметки для
                        # ключевых слов
                        logging.info("Изменяем название")
                        x = x.replace(y, '')  # оставляем только название режима
                    for i in range(0, 3):
                        x = x + pat_rez[i] if rez[i] == 1 else x  # добавляем необходимые ключевые слова
                logging.info("Записываем новый словарь")
                new_book[x] = book[name_list]
            name = [x for x in new_book.keys()]
            # Конец
            # wb = load_workbook(pathlib.Path(path, file), data_only=True)  # Откроем книгу.
            # # wb = load_workbook(path + '\\' + file, data_only=True)  # Откроем книгу.
            # book_name = str(file.rsplit('.xlsx', maxsplit=1)[0])  # Определение названия exel.
            # name = wb.sheetnames  # Список листов.
            # logging.info("Проверяем на названия файлов для ЦП")
            # for name_list in name:
            #     if re.search(r'_ЦП', name_list) or re.search(r'\.m', name_list) or re.search(r'\.v', name_list):
            #         for elem in range(0, len(name)):  # поиск и устранение неточностей в названиях вкладок ЦП
            #             if re.search(r'_ЦП', name[elem]) or re.search(r'\.m', name[elem]) or \
            #                     re.search(r'\.v', name[elem]):  # проверяем интересующие нас названия
            #                 logging.info("Нашли название" + name[elem])
            #                 rez = []
            #                 x = name[elem]
            #                 # pat = ['_ЦП', '.m', '.v']  # список ключевых слов для поиска в ЦП
            #                 # pat_rez = ['_ЦП', '.m', '.v']
            #                 for y in pat:  # прогоняем список
            #                     logging.info("Ищем совпадение в нашем списке")
            #                     if y == '.v':
            #                         replace = re.findall(r'.v\d', x)
            #                         if replace:
            #                             y = replace[0]
            #                             pat_rez[2] = y
            #                     rez.append(1) if x.find(y) != -1 else rez.append(-1)  # добавляем заметки для
            #                     # ключевых слов
            #                     logging.info("Изменяем название")
            #                     x = x.replace(y, '')  # оставляем только название режима
            #                 for i in range(0, 3):
            #                     x = x + pat_rez[i] if rez[i] == 1 else x  # добавляем необходимые ключевые слова
            #                 logging.info("Переименовываем лист")
            #                 worksheet = wb[name[elem]]  # выбираем лист с именем
            #                 worksheet.title = x  # переименовываем лист
            #         logging.info("Сохраняем книгу с новыми названиями")
            #         wb.save(filename=pathlib.Path(path, file))  # сохраняем книгу
            #         wb.close()
            #         break
            # logging.info("Открываем книгу ещё раз если закрыли её в предыдущем цикле")  # Проверить надо ли
            # wb = load_workbook(pathlib.Path(path, file), data_only=True)  # Откроем книгу.
            # name = wb.sheetnames  # Список листов.
            logging.info("Проверяем на совпадение названий с файлом описания")
            if name != mode:  # проверяем названия на соответствия
                logging.info("Названия не совпадают")
                output = 'В заказе ' + path.rpartition('\\')[2] + ' названия режимов в исходнике ' + str(file) + \
                         ' не совпадают с описанием: '
                for i_out, name_isx in enumerate(name):
                    if mode.count(name_isx) == 0:
                        output += str(i_out) + ') режим ' + str(name_isx) + '; '
                error.append(output.strip(' '))
            else:
                for sheet in name:  # Загоняем в txt.
                    line_doing.emit(f'Проверяем режимы в {file} на ошибки ({now_doc} из {all_doc})')
                    logging.info("Проверяем документы на наличие ошибок")
                    if sheet.lower() != 'описание':
                        # df = pd.read_excel(pathlib.Path(path, file), sheet_name=sheet, header=None)
                        df = new_book[sheet]
                        logging.info("Смотрим есть ли ошибки")
                        if twelve_sectors:
                            alphabet = [chr(i) for i in range(65, 90)]
                            df = df.fillna(0.0000001)
                            for column in df.columns:
                                data = df[column]
                                try:  # Блок try для отлова текста в значениях
                                    if data.astype(float).all():
                                        continue
                                    else:
                                        error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' +
                                                     file + ' в режиме '
                                                     + sheet + ' в колонке ' + str(column + 1) +
                                                     ' неведомая штука (не преобразовывается ни в строку,'
                                                     ' ни в значение)!')
                                except ValueError:
                                    for i, row in enumerate(df[column]):
                                        if type(row) == str:
                                            error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' +
                                                         file + ' в режиме '
                                                         + sheet + ' в ячейке ' + alphabet[column] + str(i + 1) +
                                                         ' есть текстовое значение!')
                        else:
                            df = df.fillna(False)
                            for i, row in enumerate(df.itertuples(index=False)):
                                try:  # Try/except блок для отлова листов с надписью «не обнаружено»
                                    frq, s, n = row[0], row[1], row[2]
                                    # if type(frq) is str:
                                    if isinstance(frq, str):
                                        # frq = float(frq.replace(',', '.'))
                                        error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' + file +
                                                     ' в режиме ' + sheet +
                                                     ' в строке ' + str(i + 1) + ' записано текстовое значение!')
                                    if s:
                                        if isinstance(s, float) or isinstance(s, int):
                                            if n is False:
                                                error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' +
                                                             file + ' в режиме '
                                                             + sheet + ' на частоте ' + str(round(frq, 4)) +
                                                             ' есть значение сигнала, но нет шума!')
                                        else:
                                            error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' +
                                                         file + ' в режиме '
                                                         + sheet + ' на частоте ' + str(round(frq, 4)) +
                                                         ' сигнал указан как текстовое значение')
                                    if n:
                                        if isinstance(n, float) or isinstance(n, int):
                                            if s is False:
                                                error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' +
                                                             file + ' в режиме ' +
                                                             sheet + ' на частоте ' + str(round(frq, 4)) +
                                                             ' есть значение шума, но нет сигнала!')
                                        else:
                                            error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' +
                                                         file + ' в режиме '
                                                         + sheet + ' на частоте ' + str(round(frq, 4)) +
                                                         ' шум указан как текстовое значение')
                                    if (s and (isinstance(s, float) or isinstance(s, int))) and\
                                            (n and (isinstance(n, float) or isinstance(n, int))) and\
                                            (no_freq_lim is False):
                                        if s < n:
                                            error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' +
                                                         file + ' в режиме ' +
                                                         sheet + ' на частоте ' +
                                                         str(round(frq, 4)) + ' значения шума больше сигнала!')
                                        elif s == n:
                                            error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' +
                                                         file + ' в режиме ' +
                                                         sheet + ' на частоте ' +
                                                         str(round(frq, 4)) + ' одинаковые значения сигнала и шума!')
                                        elif s > 100:
                                            error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' +
                                                         file + ' в режиме ' +
                                                         sheet + ' на частоте ' +
                                                         str(round(frq, 4)) + ' слишком большое значение сигнала!')
                                except IndexError:
                                    pass
            cp += per
            if error:
                logging.info("Добавляем ошибки")
                for e in error:
                    output_error.append(e)
                # wb.close()
                line_progress.emit(f'Выполнено {int(cp)} %')
                progress.emit(int(cp))
                continue
            line_doing.emit(f'Создаем txt файлы для документа {file} ({now_doc} из {all_doc})')
            logging.info("Ошибок нет, записываем в txt")
            logging.info("Создаем папку для txt файлов")
            if os.path.exists(pathlib.Path(path, 'txt', book_name)) is False:
                os.makedirs(pathlib.Path(path, 'txt', book_name))
                os.chdir(pathlib.Path(path, 'txt', book_name))
                for sheet in name:
                    if re.findall(r'_lin', sheet) or re.findall(r'_linux', sheet):
                        name_sheet = sheet.upper()
                    elif re.findall(r'_win', sheet) or re.findall(r'_windows', sheet):
                        name_sheet = sheet.lower()
                    else:
                        name_sheet = sheet
                    df = new_book[sheet]
                    # df = pd.read_excel(pathlib.Path(path, file), sheet_name=sheet, header=None)
                    if df.empty or type(df.iloc[0, 0]) == str:
                        with open(pathlib.Path(path, 'txt', book_name, name_sheet + '.txt'), 'w'):
                            pass
                    else:
                        if sheet.lower() != 'описание':
                            if twelve_sectors is False:
                                df = df.drop(df.columns[[i for i in df.columns.tolist() if i > 2]], axis=1)
                                if [0, 1, 2] in df.columns.tolist():
                                    df = df[[0, 1, 2]]
                                df = df.dropna()
                            else:
                                df = df.fillna(0)
                        df = df.round(4)
                        df.to_csv(pathlib.Path(path, 'txt', book_name, name_sheet + '.txt'),
                                  index=None, sep='\t', mode='w', header=None)
            # wb.close()
            line_progress.emit(f'Выполнено {int(cp)} %')
            progress.emit(int(cp))
        return {'error': output_error, 'cp': cp, 'now_doc': now_doc, 'cancel': False, 'base_exception': False}
    # Подумать что тут с исключениями
    except BaseException as es:
        return {'base_exception': True, 'text': es, 'trace': traceback.format_exc()}

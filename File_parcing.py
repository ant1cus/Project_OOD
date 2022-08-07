import os
import threading
import pandas as pd
import re
import traceback
from PyQt5.QtCore import QThread, pyqtSignal
from openpyxl import load_workbook


class FileParcing(QThread):
    progress = pyqtSignal(int)  # Сигнал для прогресс бара
    status = pyqtSignal(str)  # Сигнал для статус бара
    messageChanged = pyqtSignal(str, str)
    errors = pyqtSignal()

    def __init__(self, output):  # Список переданных элементов.
        QThread.__init__(self)
        self.path = output[0]
        self.group_file = output[1]
        self.logging = output[2]
        self.q = output[3]
        self.event = threading.Event()

    def run(self):
        try:
            progress = 0
            self.logging.info("Начинаем")
            self.status.emit('Старт')
            self.progress.emit(progress)
            list_file = os.listdir(self.path)
            # Сохраним нужное нам описание режимов.
            self.logging.info("Читаем txt и сохраняем режимы")
            txt_files = filter(lambda x: x.endswith('.txt'), list_file)
            for file in sorted(txt_files):
                try:
                    with open(self.path + '\\' + file, mode='r', encoding="utf-8-sig") as f:
                        self.logging.info("Кодировка utf-8-sig")
                        mode_1 = f.readlines()
                        mode_1 = [line.rstrip() for line in mode_1]
                except UnicodeDecodeError:
                    with open(self.path + '\\' + file, mode='r') as f:
                        self.logging.info("Другая кодировка")
                        mode_1 = f.readlines()
                        mode_1 = [line.rstrip() for line in mode_1]
            mode = [x for x in mode_1 if x]
            try:
                os.makedirs(self.path + '\\txt\\')
                self.logging.info("Создаем папку txt")
            except (FileExistsError, AttributeError):
                self.logging.info("Запоминаем какие папки уже есть внутри папки txt")
                parcing_file = os.listdir(self.path + '\\txt')
            # Работа с исходниками.
            # Отсортируем нужные нам файлы xlsx.
            exel_files = filter(lambda x: x.endswith('.xlsx') and ('~' not in x) and (x[:-4] not in parcing_file),
                                list_file)
            errors = []
            self.logging.info("Начинаем прохождение по файлам excel")
            for file in sorted(exel_files):
                self.logging.info("Открываем книгу")
                wb = load_workbook(self.path + '\\' + file, data_only=True)  # Откроем книгу.
                book_name = str(file.rsplit('.xlsx', maxsplit=1)[0])  # Определение названия exel.
                name = wb.sheetnames  # Список листов.
                pat = ['_ЦП', '.m', '.v']  # список ключевых слов для поиска в ЦП
                pat_rez = ['_ЦП', '.m', '.v']
                self.logging.info("Проверяем на названия файлов для ЦП")
                for name_list in name:
                    if re.search(r'_ЦП', name_list) or re.search(r'\.m', name_list) or re.search(r'\.v', name_list):
                        for elem in range(0, len(name)):  # поиск и устранение неточностей в названиях вкладок ЦП
                            if re.search(r'_ЦП', name[elem]) or re.search(r'\.m', name[elem]) or \
                                    re.search(r'\.v', name[elem]):  # проверяем интересующие нас названия
                                self.logging.info("Нашли название" + name[elem])
                                rez = []
                                x = name[elem]
                                for y in pat:  # прогоняем список
                                    self.logging.info("Ищем совпадение в нашем списке")
                                    if y == '.v':
                                        replace = re.findall(r'.v\d', x)
                                        if replace:
                                            y = replace[0]
                                            pat_rez[2] = y
                                    rez.append(1) if x.find(y) != -1 else rez.append(-1)  # добавляем заметки для
                                    # ключевых слов
                                    self.logging.info("Изменяем название")
                                    x = x.replace(y, '')  # оставляем только название режима
                                for i in range(0, 3):
                                    x = x + pat_rez[i] if rez[i] == 1 else x  # добавляем необходимые ключевые слова
                                self.logging.info("Переименовываем лист")
                                worksheet = wb[name[elem]]  # выбираем лист с именем
                                worksheet.title = x  # переименовываем лист
                        self.logging.info("Сохраняем книгу с новыми названиями")
                        wb.save(filename=file)  # сохраняем книгу
                        wb.close()
                        break
                self.logging.info("Открываем книгу ещё раз если закрыли её в предыдущем цикле")  # Проверить надо ли
                wb = load_workbook(self.path + '\\' + file, data_only=True)  # Откроем книгу.
                name = wb.sheetnames  # Список листов.
                self.logging.info("Проверяем на совпадение названий с файлом описания")
                error = []
                if name != mode:  # проверяем названия на соответствия
                    self.logging.info("Названия не совпадают")
                    errors.append('Названия режимов в исходнике' + str(file) + ' не совпадают с описанием:\n')
                    for name_isx in name:
                        if mode.count(name_isx) == 0:
                            errors.append('режим ' + str(name_isx) + '\n')
                else:
                    self.logging.info("Создаем папку для txt файлов")
                    os.makedirs(self.path + '\\txt\\' + book_name)
                    os.chdir(self.path + "\\txt\\" + book_name)
                    for sheet in name:  # Загоняем в txt.
                        self.logging.info("Проверяем документы на наличие ошибок")
                        if sheet.lower() != 'описание':
                            df = pd.read_excel(self.path + '\\' + file, sheet_name=sheet, header=None)
                            df = df.fillna(False)
                            self.logging.info("Смотрим есть ли ошибки")
                            for i, row in enumerate(df.itertuples(index=False)):
                                try:  # Try/except блок для отлова листов с надписью «не обнаружено»
                                    frq, s, n = row[0], row[1], row[2]
                                    if type(frq) is str:
                                        error.append('В исходнике ' + file + ' в режиме ' + sheet + ' в строке ' +
                                                     str(i) + ' записано текстовое значение!\n')
                                    else:
                                        frq = int(row[0]) if str(row[0]).find('.') == -1 else\
                                            "%.4f" % float(frq)
                                    if (type(s) is float or type(s) is int) and n is False:
                                        error.append('В исходнике ' + file + ' в режиме ' + sheet + ' на частоте '
                                                     + str(row[0]) + ' есть значение сигнала, но нет шума!\n')
                                    elif (type(n) is float or type(n) is int) and s is False:
                                        error.append('В исходнике ' + file + ' в режиме ' + sheet + ' на частоте '
                                                     + str(row[0]) + ' есть значение шума, но нет сигнала!\n')
                                    elif (type(s) is float or type(s) is int) and (type(n) is float or type(n) is int):
                                        s = round(float(row[1]), 2)
                                        n = round(float(row[2]), 2)
                                        if s < n:
                                            error.append('В исходнике ' + file + ' в режиме ' + sheet + ' на частоте ' +
                                                         str(row[0]) + ' значения шума больше сигнала!\n')
                                        elif s == n:
                                            error.append('В исходнике ' + file + ' в режиме ' + sheet + ' на частоте ' +
                                                         str(row[0]) + ' одинаковые значения сигнала и шума!\n')
                                        elif s > 100:
                                            error.append('В исходнике ' + file + ' в режиме ' + sheet + ' на частоте ' +
                                                         str(row[0]) + ' слишком большое значение сигнала!\n')
                                    for j, el in enumerate([frq, s, n]):
                                        df.iloc[i, j] = el
                                except IndexError:
                                    pass
                if error is False:
                    self.logging.info("Ошибок нет, записываем в txt")
                    for sheet in name:
                        df = pd.read_excel(self.path + '\\' + file, sheet_name=sheet, header=None)
                        df = df.dropna()
                        df.to_csv(self.path + '\\txt\\' + book_name + '\\' + sheet + '.txt',
                                  index=None, sep='\t', mode='w', header=None)
                else:
                    self.logging.info("Добавляем ошибки")
                    errors.append(error)
                wb.close()
            if errors:
                self.logging.info("Выводим ошибки")
                self.q.put(errors)
                self.errors.emit()
                self.status.emit('В файлах присутствуют ошибки')
            else:
                self.logging.info("Конец работы программы")
                self.status.emit('Готово')
            return
        except BaseException as es:
            self.logging.error(es)
            self.logging.error(traceback.format_exc())
            self.progress.emit(0)
            self.status.emit('Ошибка!')
            os.chdir('C://')
            return


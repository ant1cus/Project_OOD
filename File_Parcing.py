import os
import threading
import pandas as pd
import re
import traceback
from PyQt5.QtCore import QThread, pyqtSignal
from openpyxl import load_workbook


class FileParcing(QThread):
    progress = pyqtSignal(int)  # Сигнал для прогресс бара
    status = pyqtSignal(str)  # Сигнал для статус бара
    messageChanged = pyqtSignal(str, str)
    errors = pyqtSignal()

    def __init__(self, output):  # Список переданных элементов.
        QThread.__init__(self)
        self.path = output[0]
        self.all_progress = output[1]
        self.group_file = output[2]
        self.logging = output[3]
        self.q = output[4]
        self.event = threading.Event()

    def run(self):
        def file_parcing(path, logging, status, progress, per, cp):
            list_file = os.listdir(path)
            # Сохраним нужное нам описание режимов.
            logging.info("Читаем txt и сохраняем режимы для " + path)
            status.emit('Считываем режимы из текстового файла для заказа ' + path.rpartition('\\')[2])
            txt_files = filter(lambda x: x.endswith('.txt'), list_file)
            for file in sorted(txt_files):
                try:
                    with open(path + '\\' + file, mode='r', encoding="utf-8-sig") as f:
                        logging.info("Кодировка utf-8-sig")
                        mode_1 = f.readlines()
                        mode_1 = [line.rstrip() for line in mode_1]
                except UnicodeDecodeError:
                    with open(path + '\\' + file, mode='r') as f:
                        logging.info("Другая кодировка")
                        mode_1 = f.readlines()
                        mode_1 = [line.rstrip() for line in mode_1]
            mode = [x for x in mode_1 if x]
            parcing_file = []
            if os.path.exists(path + '\\txt'):
                logging.info("Запоминаем какие папки уже есть внутри папки txt")
                parcing_file = os.listdir(path + '\\txt')
            # Работа с исходниками.
            # Отсортируем нужные нам файлы xlsx.
            exel_files = filter(lambda x: x.endswith('.xlsx') and ('~' not in x) and (x[:-4] not in parcing_file),
                                list_file)
            logging.info("Начинаем прохождение по файлам excel")
            output_error = []
            for file in sorted(exel_files):
                status.emit('Проверяем названия рабочих листов в документе ' + file)
                error = []
                logging.info("Открываем книгу")
                wb = load_workbook(path + '\\' + file, data_only=True)  # Откроем книгу.
                book_name = str(file.rsplit('.xlsx', maxsplit=1)[0])  # Определение названия exel.
                name = wb.sheetnames  # Список листов.
                pat = ['_ЦП', '.m', '.v']  # список ключевых слов для поиска в ЦП
                pat_rez = ['_ЦП', '.m', '.v']
                logging.info("Проверяем на названия файлов для ЦП")
                for name_list in name:
                    if re.search(r'_ЦП', name_list) or re.search(r'\.m', name_list) or re.search(r'\.v', name_list):
                        for elem in range(0, len(name)):  # поиск и устранение неточностей в названиях вкладок ЦП
                            if re.search(r'_ЦП', name[elem]) or re.search(r'\.m', name[elem]) or \
                                    re.search(r'\.v', name[elem]):  # проверяем интересующие нас названия
                                logging.info("Нашли название" + name[elem])
                                rez = []
                                x = name[elem]
                                for y in pat:  # прогоняем список
                                    logging.info("Ищем совпадение в нашем списке")
                                    if y == '.v':
                                        replace = re.findall(r'.v\d', x)
                                        if replace:
                                            y = replace[0]
                                            pat_rez[2] = y
                                    rez.append(1) if x.find(y) != -1 else rez.append(-1)  # добавляем заметки для
                                    # ключевых слов
                                    logging.info("Изменяем название")
                                    x = x.replace(y, '')  # оставляем только название режима
                                for i in range(0, 3):
                                    x = x + pat_rez[i] if rez[i] == 1 else x  # добавляем необходимые ключевые слова
                                logging.info("Переименовываем лист")
                                worksheet = wb[name[elem]]  # выбираем лист с именем
                                worksheet.title = x  # переименовываем лист
                        logging.info("Сохраняем книгу с новыми названиями")
                        wb.save(filename=file)  # сохраняем книгу
                        wb.close()
                        break
                logging.info("Открываем книгу ещё раз если закрыли её в предыдущем цикле")  # Проверить надо ли
                wb = load_workbook(path + '\\' + file, data_only=True)  # Откроем книгу.
                name = wb.sheetnames  # Список листов.
                logging.info("Проверяем на совпадение названий с файлом описания")
                if name != mode:  # проверяем названия на соответствия
                    logging.info("Названия не совпадают")
                    output = 'В заказе ' + path.rpartition('\\')[2] + ' названия режимов в исходнике ' + str(file) +\
                             ' не совпадают с описанием: '
                    for i_out, name_isx in enumerate(name):
                        if mode.count(name_isx) == 0:
                            output += str(i_out) + ') режим ' + str(name_isx) + '; '
                    error.append(output.strip(' '))
                else:
                    for sheet in name:  # Загоняем в txt.
                        status.emit('Проверяем рабочие листы в документе ' + file + ' на правильность заполнения')
                        logging.info("Проверяем документы на наличие ошибок")
                        if sheet.lower() != 'описание':
                            df = pd.read_excel(path + '\\' + file, sheet_name=sheet, header=None)
                            df = df.fillna(False)
                            logging.info("Смотрим есть ли ошибки")
                            for i, row in enumerate(df.itertuples(index=False)):
                                try:  # Try/except блок для отлова листов с надписью «не обнаружено»
                                    frq, s, n = row[0], row[1], row[2]
                                    if type(frq) is str:
                                        error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' + file +
                                                     ' в режиме ' + sheet +
                                                     ' в строке ' + str(i) + ' записано текстовое значение!')
                                    if s:
                                        if type(s) is float or type(s) is int:
                                            if n is False:
                                                error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' +
                                                             file + ' в режиме '
                                                             + sheet + ' на частоте ' + str(round(frq, 4)) +
                                                             ' есть значение сигнала, но нет шума!')
                                        else:
                                            error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' +
                                                         file + ' в режиме '
                                                         + sheet + ' на частоте ' + str(round(frq, 4)) +
                                                         ' сигнал указан как текстовое значение')
                                    if n:
                                        if type(n) is float or type(n) is int:
                                            if s is False:
                                                error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' +
                                                             file + ' в режиме ' +
                                                             sheet + ' на частоте ' + str(round(frq, 4)) +
                                                             ' есть значение шума, но нет сигнала!')
                                        else:
                                            error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' +
                                                         file + ' в режиме '
                                                         + sheet + ' на частоте ' + str(round(frq, 4)) +
                                                         ' шум указан как текстовое значение')
                                    if (type(s) is float or type(s) is int) and (type(n) is float or type(n) is int):
                                        if s < n:
                                            error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' +
                                                         file + ' в режиме ' +
                                                         sheet + ' на частоте ' +
                                                         str(round(frq, 4)) + ' значения шума больше сигнала!')
                                        elif s == n:
                                            error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' +
                                                         file + ' в режиме ' +
                                                         sheet + ' на частоте ' +
                                                         str(round(frq, 4)) + ' одинаковые значения сигнала и шума!')
                                        elif s > 100:
                                            error.append('В заказе ' + path.rpartition('\\')[2] + ' в исходнике ' +
                                                         file + ' в режиме ' +
                                                         sheet + ' на частоте ' +
                                                         str(round(frq, 4)) + ' слишком большое значение сигнала!')
                                except IndexError:
                                    pass
                cp = cp + per
                if error:
                    status.emit('Добавляем ошибки для документа ' + file)
                    logging.info("Добавляем ошибки")
                    for e in error:
                        output_error.append(e)
                    wb.close()
                    progress.emit(cp)
                else:
                    status.emit('Создаем txt файлы для документа ' + file)
                    logging.info("Ошибок нет, записываем в txt")
                    logging.info("Создаем папку для txt файлов")
                    if os.path.exists(path + '\\txt\\' + book_name) is False:
                        os.makedirs(path + '\\txt\\' + book_name)
                        os.chdir(path + "\\txt\\" + book_name)
                        for sheet in name:
                            df = pd.read_excel(path + '\\' + file, sheet_name=sheet, header=None)
                            if sheet.lower() != 'описание':
                                df = df.dropna()
                            df = df.round(4)
                            df.to_csv(path + '\\txt\\' + book_name + '\\' + sheet + '.txt',
                                      index=None, sep='\t', mode='w', header=None)
                    wb.close()
                    progress.emit(cp)
            return {'error': output_error, 'cp': cp}
        try:
            current_progress = 0
            self.logging.info("Начинаем")
            self.status.emit('Старт')
            self.progress.emit(current_progress)
            percent = 100/self.all_progress
            errors = []
            succsess_path = []
            error_path = []
            if self.group_file:
                for folder in os.listdir(self.path):
                    if os.path.isdir(self.path + '\\' + folder):
                        err = file_parcing(self.path + '\\' + folder, self.logging, self.status, self.progress, percent,
                                           current_progress)
                        if err['error']:
                            error_path.append(self.path + '\\' + folder)
                            for element in err['error']:
                                errors.append(element)
                        else:
                            succsess_path.append(self.path + '\\' + folder)
                        current_progress = err['cp']
            else:
                err = file_parcing(self.path, self.logging, self.status, self.progress, percent, current_progress)
                if err['error']:
                    errors = err['error']
                    error_path.append(self.path)
                else:
                    succsess_path.append(self.path)
            if succsess_path:
                self.q.put({'Прошедшие заказы:': '\n'.join(succsess_path)})
                self.errors.emit()
            if error_path:
                self.q.put({'Заказы с ошибками:': '\n'.join(error_path)})
                self.errors.emit()
            if errors:
                self.logging.info("Выводим ошибки")
                self.q.put({'errors': errors})
                self.errors.emit()
                self.status.emit('В файлах присутствуют ошибки')
            else:
                self.logging.info("Конец работы программы")
                self.status.emit('Готово')
            os.chdir('C:\\')
            return
        except BaseException as es:
            self.logging.error(es)
            self.logging.error(traceback.format_exc())
            self.progress.emit(0)
            self.status.emit('Ошибка!')
            os.chdir('C:\\')
            return

